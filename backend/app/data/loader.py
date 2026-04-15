"""Load and validate customer account data from the sample workbook."""

import csv
from functools import lru_cache
from io import BytesIO, StringIO
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import load_workbook
from pydantic import ValidationError

from app.models.customer import CustomerAccount

BACKEND_ROOT = Path(__file__).resolve().parents[2]
DATA_RELATIVE_PATH = Path("data") / "sample_customers_q3_2025.xlsx"
DATA_FILE = BACKEND_ROOT / DATA_RELATIVE_PATH
EXPECTED_HEADERS = (
    "account_name",
    "plan_type",
    "active_users",
    "usage_growth_qoq",
    "automation_adoption_pct",
    "tickets_last_quarter",
    "avg_response_time",
    "nps_score",
    "preferred_channel",
    "scat_score",
    "risk_engine_score",
    "crm_notes",
    "feedback_summary",
)


def _clean_cell(value: object) -> object:
    if isinstance(value, str):
        return value.strip()
    return value


def parse_rows(rows: Iterable[Sequence[object]], *, source_label: str) -> list[CustomerAccount]:
    row_iterator = iter(rows)
    header_row = next(row_iterator, None)

    if header_row is None:
        raise ValueError(f"{source_label} is empty")

    normalized_headers = tuple(_clean_cell(value) for value in header_row)
    if normalized_headers != EXPECTED_HEADERS:
        raise ValueError(
            "Unexpected workbook schema. "
            f"Expected {EXPECTED_HEADERS}, got {normalized_headers}."
        )

    accounts: list[CustomerAccount] = []
    for row_number, row in enumerate(row_iterator, start=2):
        if row is None or all(value is None for value in row):
            continue
        if len(row) != len(normalized_headers):
            raise ValueError(
                f"Invalid customer row at line {row_number}: "
                f"expected {len(normalized_headers)} columns, got {len(row)}."
            )

        payload = {
            header: _clean_cell(value)
            for header, value in zip(normalized_headers, row, strict=True)
        }

        try:
            accounts.append(CustomerAccount.model_validate(payload))
        except ValidationError as exc:
            raise ValueError(f"Invalid customer row at line {row_number}") from exc

    return accounts


def parse_customer_file(file_name: str, file_bytes: bytes) -> list[CustomerAccount]:
    suffix = Path(file_name).suffix.lower()

    if suffix == ".xlsx":
        workbook = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
        worksheet = workbook[workbook.sheetnames[0]]
        return parse_rows(worksheet.iter_rows(values_only=True), source_label=file_name)

    if suffix == ".csv":
        text = file_bytes.decode("utf-8-sig")
        reader = csv.reader(StringIO(text))
        return parse_rows(reader, source_label=file_name)

    raise ValueError("Unsupported file type. Please upload an .xlsx or .csv file.")


@lru_cache(maxsize=1)
def _load_accounts() -> tuple[CustomerAccount, ...]:
    if not DATA_FILE.exists():
        raise FileNotFoundError(f"Customer sample workbook not found at {DATA_FILE}")

    workbook = load_workbook(DATA_FILE, data_only=True, read_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    return tuple(
        parse_rows(
            worksheet.iter_rows(values_only=True),
            source_label=str(DATA_RELATIVE_PATH),
        )
    )


def get_all_accounts() -> list[CustomerAccount]:
    """Return all customer accounts from the sample dataset."""

    return [account.model_copy(deep=True) for account in _load_accounts()]


def get_account_by_name(name: str) -> CustomerAccount | None:
    """Return a single account by name, matching case-insensitively."""

    normalized_name = name.strip().casefold()
    for account in _load_accounts():
        if account.account_name.casefold() == normalized_name:
            return account.model_copy(deep=True)
    return None
